import openpyxl

from pathlib import Path
def _ensure_file_exists(file):
    p = Path(file)
    if not p.exists():
        raise FileNotFoundError(f"Excel file not found: {p.resolve()}")
    return p


def getRowCount(file, Sheetname):
    file = _ensure_file_exists(file)  # <— added
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheetname]
    rows = sheet.max_row
    workbook.close()
    return rows

def getColumnCount(file, Sheetname):
    file = _ensure_file_exists(file)  # <— added
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheetname]
    cols = sheet.max_column
    workbook.close()
    return cols

def readData(file, Sheetname, rownum, columnno):
    file = _ensure_file_exists(file)  # <— added
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook[Sheetname]
    value = sheet.cell(row=rownum, column=columnno).value
    workbook.close()
    return value

def writeData(file, Sheetname, rownum, columnno, data):
    file = _ensure_file_exists(file)  # <— added
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[Sheetname]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)
    workbook.close()
